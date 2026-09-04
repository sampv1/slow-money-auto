#!/usr/bin/env python3
"""Export one or more symbols' vnstock financial statements to an Excel workbook.

WHY IT REUSES `fa.vnstock_source.fetch_symbol` RATHER THAN CALLING THE PROVIDER
ITSELF. That function is what fills `fa_vnstock_statements`, so sharing it means
the spreadsheet and the site can never disagree about what "the FA data" is --
including the parts that are easy to get subtly wrong: the three-attempt retry
(the provider fails per CALL, not per symbol), dropping NaN/inf before they
reach a cell, and counting failed calls so a transient outage is not silently
exported as a company that files no cash-flow statement.

REQUIRES ~/.venv, NOT scripts/.venv -- `vnstock_data` is a sponsor package that
is not on PyPI. See the note at the top of fa/vnstock_source.py.

    ~/.venv/bin/python export_fa_vnstock_excel.py --symbols BVH

VALUES ARE EXPORTED RAW, in the provider's own units, with the unit named on
every row. The dashboard divides by 1e9 to show tỷ đồng, which is right for a
chart axis and wrong here: the same statement carries VNĐ line items beside
percentages and plain ratios, so one blanket conversion would silently corrupt
every row that is not money. Scaling is the reader's to apply, once they can see
which rows it applies to.

CATEGORY HEADERS ARE KEPT, as blank rows. `fetch_symbol` stores only the values
(a header carries none), but the metric dictionary it returns alongside knows
about every line -- so the sheet can be rebuilt with the statement's real
structure and indentation instead of a flat list of whatever happened to have a
number in it.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

from fa.vnstock_source import PERIOD_TYPES, STATEMENTS, fetch_symbol

# Sheet name per (statement, period_type). Excel caps a name at 31 chars and
# forbids []:*?/\ -- these are short and plain, so nothing needs sanitising.
SHEET_ORDER = [(s, p) for s in STATEMENTS.values() for p in PERIOD_TYPES]


def _period_sort_key(period: str):
    """`2026-Q2` and `2026` both sort chronologically as plain strings.

    Annual periods are bare years and quarterly ones are `YYYY-Qn`, so
    lexicographic order is already chronological within a sheet -- a sheet never
    mixes the two. Split anyway so a future `2026-H1`-style period does not
    quietly sort before `2026-Q1`.
    """
    parts = str(period).split("-")
    year = parts[0]
    rest = parts[1] if len(parts) > 1 else ""
    return (year, rest)


def build_sheets(rows: list[dict], metrics: list[dict]) -> dict:
    """(statement, period_type) -> (periods, [(metric, {period: value})])."""
    by_id = {m["metric_id"]: m for m in metrics}
    out = {}
    for stmt, ptype in SHEET_ORDER:
        sel = [r for r in rows if r["statement"] == stmt and r["period_type"] == ptype]
        if not sel:
            continue
        periods = sorted({r["period"] for r in sel}, key=_period_sort_key)

        # Every metric the DICTIONARY knows for this statement, not just the
        # ones carrying a value -- that is what keeps the headers and the
        # indentation. Ordered by the provider's own `display_order` so the
        # sheet reads in the order the statement is filed.
        wanted = [m for m in metrics if m["statement"] == stmt]
        wanted.sort(key=lambda m: (m.get("display_order") is None,
                                   m.get("display_order") or 0,
                                   m["metric_id"]))
        values: dict[str, dict] = {}
        for r in sel:
            for mid, v in r["items"].items():
                values.setdefault(mid, {})[r["period"]] = v

        # A metric with no value in ANY period of this sheet is dropped: an
        # all-blank row is not structure, it is a line the provider files on the
        # other period grain.
        lines = [(by_id.get(m["metric_id"], m), values.get(m["metric_id"], {}))
                 for m in wanted
                 if values.get(m["metric_id"]) or _is_header(m, values)]
        if lines:
            out[(stmt, ptype)] = (periods, lines)
    return out


def _is_header(metric: dict, values: dict) -> bool:
    """A level-1 line with no value of its own, kept because it names a block."""
    return (metric.get("level") or 99) <= 1 and metric["metric_id"] not in values


def write_workbook(path: Path, symbol: str, sheets: dict, metrics: list[dict],
                   failed_calls: int) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    info = wb.create_sheet("info")
    info.append(["Symbol", symbol])
    info.append(["Source", "vnstock_data (sponsor API), via fa/vnstock_source.py"])
    info.append(["Exported", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    info.append(["Failed provider calls", failed_calls])
    if failed_calls:
        info.append(["WARNING",
                     "A statement below may be short or missing because a call "
                     "failed, not because the company does not file it. Re-run."])
    info.append([])
    info.append(["Sheet", "Statement", "Period type", "Lines", "Periods", "From", "To"])
    for (stmt, ptype), (periods, lines) in sheets.items():
        info.append([f"{stmt}_{ptype}", stmt, ptype, len(lines), len(periods),
                     periods[0], periods[-1]])
    for c in info[1]:
        c.font = Font(bold=True)
    for c in info[7]:
        c.font = Font(bold=True)
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 52

    bold = Font(bold=True)
    for (stmt, ptype), (periods, lines) in sheets.items():
        ws = wb.create_sheet(f"{stmt}_{ptype}")
        header = ["metric_id", "name_vi", "unit", "level"] + periods
        ws.append(header)
        for c in ws[1]:
            c.font = bold
            c.alignment = Alignment(horizontal="center", vertical="bottom")

        for metric, vals in lines:
            level = metric.get("level") or 1
            name = metric.get("name_vi") or ""
            row = [metric["metric_id"],
                   ("    " * max(0, level - 1)) + name,
                   metric.get("unit"),
                   level]
            row += [vals.get(p) for p in periods]
            ws.append(row)
            # A block header carries no figures; bolding it is what makes the
            # indentation readable as a hierarchy rather than as ragged text.
            if not vals:
                for c in ws[ws.max_row]:
                    c.font = bold

        for r in ws.iter_rows(min_row=2, min_col=5):
            for c in r:
                if isinstance(c.value, (int, float)):
                    c.number_format = "#,##0" if abs(c.value) >= 1000 else "0.0000"

        ws.freeze_panes = "E2"
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 58
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 6
        for i in range(len(periods)):
            ws.column_dimensions[get_column_letter(5 + i)].width = 16

    md = wb.create_sheet("metric_dictionary")
    md.append(["metric_id", "statement", "name_vi", "unit", "display_order", "level"])
    for c in md[1]:
        c.font = bold
    for m in sorted(metrics, key=lambda x: (x["statement"],
                                            x.get("display_order") is None,
                                            x.get("display_order") or 0)):
        md.append([m["metric_id"], m["statement"], m.get("name_vi"), m.get("unit"),
                   m.get("display_order"), m.get("level")])
    md.freeze_panes = "A2"
    md.column_dimensions["A"].width = 42
    md.column_dimensions["C"].width = 58

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--symbols", default="BVH",
                    help="comma-separated tickers (default: BVH)")
    ap.add_argument("--out-dir", default="../tmp",
                    help="directory for the workbook(s) (default: ../tmp)")
    args = ap.parse_args()

    symbols = [s.strip().upper() for s in args.symbols.split(",") if s.strip()]
    stamp = datetime.now().strftime("%Y%m%d")
    rc = 0

    for symbol in symbols:
        print(f"[{symbol}] fetching ...", flush=True)
        rows, metrics, failed = fetch_symbol(symbol)
        if not rows:
            print(f"[{symbol}] NOTHING RETURNED "
                  f"({failed} failed call(s)) -- nothing written")
            rc = 1
            continue

        sheets = build_sheets(rows, metrics)
        out = Path(args.out_dir) / f"fa_vnstock_{symbol}_{stamp}.xlsx"
        write_workbook(out, symbol, sheets, metrics, failed)

        total = sum(len(v[1]) for v in sheets.values())
        print(f"[{symbol}] {len(sheets)} sheets, {total} lines, "
              f"{len(metrics)} metrics -> {out}")
        # Loud, not fatal: the file is still useful, but a short statement in it
        # is a provider failure rather than a company that does not file one.
        if failed:
            print(f"[{symbol}] WARNING: {failed} provider call(s) failed after "
                  f"retries -- a statement may be missing. Re-run to fill it.")
            rc = 1

    return rc


if __name__ == "__main__":
    sys.exit(main())
