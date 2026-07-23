"""Parse the FA Excel exports into normalized rows for the DB.

Two workbooks (both in data/ by default):

  Data_FiinPro.xlsx  — multi-sheet financials. Each sheet has a 6-row preamble,
                       the header on row 7 (0-indexed: row index 7), data from
                       row 8. ID columns A-D = STT, Mã, Tên, Sàn. Metric columns
                       carry a quarter in their header (3 label formats). Sheets:
                         EPS                 -> single-quarter EPS ("EPS Qn/YYYY")
                         Biên lãi gộp + ròng -> gross & net margin ("Biên lãi … Qn/YYYY")
                         Doanh thu           -> single-quarter net revenue
                         Nợ ngắn hạn         -> short-term financial debt
                         Nợ dài hạn          -> long-term financial debt
                         Vốn chủ             -> total owner's equity
                         ROE                 -> ROE % (TTM)
                       Sheets are joined by symbol into one row per (symbol, period).

  PE.xlsx            — annual P/E (Chỉ số năm) per symbol, 2021-2025. Header on
                       row 0; ID columns A-D; metric columns carry a year.

Everything is an additive upsert of only the rows present — partial files are
expected and re-importable.
"""

import re
import unicodedata

import openpyxl

HEADER_ROW_IDX = 7   # Data_FiinPro: 0-indexed header row
SYMBOL_COL = 1       # column B

# Quarter labels appear in 3 forms across sheets:
#   "… Quý: Q2 ⏎ Năm: 2024 …"      (statement line items)
#   "… Quý: Q2.2024 …"             (TTM ratios, shares)
#   "EPS Q2/2024" / "Biên lãi gộp Q2/2024"  (pre-computed per-quarter metrics)
_RE_Q_NAM = re.compile(r"Quý:\s*Q([1-4]).*?Năm:\s*(\d{4})", re.DOTALL)
_RE_Q_DOT = re.compile(r"Quý:\s*Q([1-4])\.(\d{4})")
_RE_Q_SLASH = re.compile(r"Q([1-4])/(\d{4})")


def _parse_quarter(header: str):
    """Return 'YYYY-Qn' from a header cell, or None if no quarter is present."""
    if not header:
        return None
    for rx in (_RE_Q_NAM, _RE_Q_DOT, _RE_Q_SLASH):
        m = rx.search(str(header))
        if m:
            return f"{m.group(2)}-Q{m.group(1)}"
    return None


def _num(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return None if f != f else f  # drop NaN


def _sheet_quarter_map(ws, header_row_idx: int, want_prefix: str | None = None):
    """Map column index -> 'YYYY-Qn' for metric columns of one sheet.

    `want_prefix`: when a sheet holds several blocks (e.g. the margin sheet has
    revenue + gross profit + net profit + gross margin + net margin), restrict to
    columns whose header starts with this text so we pick the right block.
    """
    header = None
    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 1, values_only=True)):
        header = row
        break
    out = {}
    for idx, cell in enumerate(header or []):
        if idx < 4 or cell is None:
            continue
        text = str(cell)
        if want_prefix and not text.strip().startswith(want_prefix):
            continue
        q = _parse_quarter(text)
        if q:
            out[idx] = q
    return out


def _read_sheet_metric(ws, header_row_idx: int, want_prefix: str | None = None):
    """Return {symbol: {period: value}} for the matching metric block of a sheet."""
    colmap = _sheet_quarter_map(ws, header_row_idx, want_prefix)
    data: dict[str, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        sym = row[SYMBOL_COL] if len(row) > SYMBOL_COL else None
        if not sym:
            continue
        sym = str(sym).strip().upper()
        per: dict[str, float] = {}
        for idx, period in colmap.items():
            if idx < len(row):
                val = _num(row[idx])
                if val is not None:
                    per[period] = val
        if per:
            data[sym] = per
    return data


# --- Single-sheet "DE - Doanh nghiệp" template (raw statement items) ---------
#
# A newer FiinProX export puts every metric in ONE sheet, one row per symbol,
# as RAW items (parent profit, shares, gross/net profit) instead of the
# pre-computed EPS / margin% columns. We match columns by (accent-stripped)
# label and DERIVE the three missing metrics — reproducing FiinProX's own values
# exactly: eps = parent profit / shares, gross_margin = gross profit / revenue,
# net_margin = net profit / revenue (margins are fractions — no ×100).
# Mirror: dashboard/src/lib/fa-import.ts.

_LEGACY_SHEETS = ("EPS", "Biên lãi gộp + ròng", "Doanh thu", "Nợ ngắn hạn",
                  "Nợ dài hạn", "Vốn chủ", "ROE")

# Direct metric label tests (pre-computed columns), on the normalized header.
_DIRECT_TESTS = {
    "eps": lambda h: "eps" in h,
    "gross_margin": lambda h: "bien lai gop" in h,
    "net_margin": lambda h: "bien lai rong" in h,
    "revenue": lambda h: "doanh thu thuan" in h,
    "st_debt": lambda h: "vay" in h and "ngan han" in h,
    "lt_debt": lambda h: "vay" in h and "dai han" in h,
    "total_equity": lambda h: "von chu so huu" in h,
    "roe_ttm": lambda h: "roe" in h,
}
# Raw statement components (not persisted) → used only to derive eps/margins.
_RAW_TESTS = {
    "parent_profit": lambda h: "loi nhuan sau thue" in h and "chu so huu" in h,
    "shares": lambda h: "cp luu hanh" in h,
    "gross_profit": lambda h: "loi nhuan gop" in h,
    "net_profit": lambda h: "loi nhuan sau thue" in h and "thu nhap doanh nghiep" in h,
}


def _norm(s) -> str:
    """Lowercase, strip Vietnamese diacritics, collapse whitespace (mirror of the TS norm())."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d").replace("Đ", "d")
    return " ".join(s.lower().split())


def _find_header_row(ws, max_scan: int = 20):
    """Locate (0-based header row index, symbol column) by finding the 'Mã' cell."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        for c, cell in enumerate(row):
            if _norm(cell) == "ma":
                return i, c
    return None


def _parse_single_sheet(ws) -> list[dict]:
    """Parse one raw-items sheet → fa_quarterly row dicts, deriving eps/margins."""
    found = _find_header_row(ws)
    if not found:
        return []
    hdr_idx, sym_col = found
    header = next(ws.iter_rows(min_row=hdr_idx + 1, max_row=hdr_idx + 1, values_only=True))

    def build_cols(tests):
        m = {}
        for field, test in tests.items():
            cols = []
            for idx, cell in enumerate(header):
                if idx <= sym_col or cell is None:
                    continue
                period = _parse_quarter(str(cell))
                if period and test(_norm(cell)):
                    cols.append((idx, period))
            if cols:
                m[field] = cols
        return m

    direct_cols = build_cols(_DIRECT_TESTS)
    raw_cols = build_cols(_RAW_TESTS)

    rows: dict[tuple[str, str], dict] = {}
    raw_by_key: dict[tuple[str, str], dict] = {}
    for row in ws.iter_rows(min_row=hdr_idx + 2, values_only=True):
        sym = row[sym_col] if len(row) > sym_col else None
        if not sym:
            continue
        sym = str(sym).strip().upper()
        for field, cols in direct_cols.items():
            for idx, period in cols:
                if idx < len(row):
                    val = _num(row[idx])
                    if val is not None:
                        key = (sym, period)
                        r = rows.get(key)
                        if r is None:
                            year, q = period.split("-Q")
                            r = {"symbol": sym, "period": period, "year": int(year), "quarter": int(q)}
                            rows[key] = r
                        r[field] = val
        for field, cols in raw_cols.items():
            for idx, period in cols:
                if idx < len(row):
                    val = _num(row[idx])
                    if val is not None:
                        raw_by_key.setdefault((sym, period), {})[field] = val

    # Derivation: pre-computed metric wins; else compute from raw components
    # (guard denominators). Create a row only when something is actually derived.
    for key, rc in raw_by_key.items():
        r = rows.get(key)
        revenue = r.get("revenue") if r else None
        derived = {}
        if (r is None or r.get("eps") is None) and rc.get("parent_profit") is not None and rc.get("shares"):
            derived["eps"] = rc["parent_profit"] / rc["shares"]
        if (r is None or r.get("gross_margin") is None) and rc.get("gross_profit") is not None and revenue:
            derived["gross_margin"] = rc["gross_profit"] / revenue
        if (r is None or r.get("net_margin") is None) and rc.get("net_profit") is not None and revenue:
            derived["net_margin"] = rc["net_profit"] / revenue
        if not derived:
            continue
        if r is None:
            sym, period = key
            year, q = period.split("-Q")
            r = {"symbol": sym, "period": period, "year": int(year), "quarter": int(q)}
            rows[key] = r
        r.update(derived)
    return list(rows.values())


def parse_fiinpro(path: str) -> list[dict]:
    """Parse a FiinProX financials workbook → fa_quarterly row dicts (present cells only).

    Dispatches by layout: the legacy multi-sheet workbook (Data_FiinPro.xlsx) uses
    the per-metric sheets below; any other workbook is treated as the single-sheet
    raw-items template and parsed with column-label matching + derivation.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if not all(s in wb.sheetnames for s in _LEGACY_SHEETS):
        for name in wb.sheetnames:
            rows = _parse_single_sheet(wb[name])
            if rows:
                return rows
        return []

    eps = _read_sheet_metric(wb["EPS"], HEADER_ROW_IDX, want_prefix="EPS ")
    gm = _read_sheet_metric(wb["Biên lãi gộp + ròng"], HEADER_ROW_IDX, want_prefix="Biên lãi gộp")
    nm = _read_sheet_metric(wb["Biên lãi gộp + ròng"], HEADER_ROW_IDX, want_prefix="Biên lãi ròng")
    rev = _read_sheet_metric(wb["Doanh thu"], HEADER_ROW_IDX, want_prefix="3. Doanh thu")
    std = _read_sheet_metric(wb["Nợ ngắn hạn"], HEADER_ROW_IDX, want_prefix="1.10")
    ltd = _read_sheet_metric(wb["Nợ dài hạn"], HEADER_ROW_IDX, want_prefix="2.8")
    eq = _read_sheet_metric(wb["Vốn chủ"], HEADER_ROW_IDX, want_prefix="II. VỐN CHỦ")
    roe = _read_sheet_metric(wb["ROE"], HEADER_ROW_IDX, want_prefix="ROE")

    blocks = {
        "eps": eps, "gross_margin": gm, "net_margin": nm, "revenue": rev,
        "st_debt": std, "lt_debt": ltd, "total_equity": eq, "roe_ttm": roe,
    }

    # union of all (symbol, period) keys that appear in any block
    rows: dict[tuple[str, str], dict] = {}
    for field, data in blocks.items():
        for sym, per in data.items():
            for period, val in per.items():
                key = (sym, period)
                r = rows.get(key)
                if r is None:
                    year, q = period.split("-Q")
                    r = {"symbol": sym, "period": period, "year": int(year), "quarter": int(q)}
                    rows[key] = r
                r[field] = val
    return list(rows.values())


def parse_pe(path: str) -> list[dict]:
    """Parse PE.xlsx → list of fa_annual_pe row dicts (only present cells)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
    # header on row 0; year columns carry "Năm: YYYY"
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    colyear = {}
    for idx, cell in enumerate(header):
        if idx < 4 or cell is None:
            continue
        m = re.search(r"Năm:\s*(\d{4})", str(cell))
        if m:
            colyear[idx] = int(m.group(1))
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sym = row[SYMBOL_COL] if len(row) > SYMBOL_COL else None
        if not sym:
            continue
        sym = str(sym).strip().upper()
        for idx, year in colyear.items():
            if idx < len(row):
                val = _num(row[idx])
                if val is not None:
                    out.append({"symbol": sym, "year": year, "pe": val})
    return out
