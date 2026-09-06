#!/usr/bin/env python3
"""
export_securities_debug.py — debug workbook for the securities (CTCK) rubric.

Built for the BA team's V10 investigation. It exports the CURRENT stored state
without repairing anything, because the point is to locate a cause rather than
to produce a healthy-looking file: C19 is N/A for the whole universe right now,
and a run that quietly recomputed it would destroy the evidence.

Two independent things are visible in here and they are NOT the same problem:

  1. C20 is N/A BY DESIGN (V9). Its formula scored 0 for every broker, so it was
     withdrawn and its 12 points leave the denominator.
  2. C19 is N/A BY REGRESSION. `run_backfill` calls `valuation_inputs` without
     the `history` argument, which defaults to None, so the core-P/E history is
     empty and the percentile cannot be formed. The daily path passes it. Sheet
     `6_Root_cause` computes both ways side by side to show the divergence.

Usage:
  python3 export_securities_debug.py                       # all 42 brokers
  python3 export_securities_debug.py --symbols VCK SSI HBS
  python3 export_securities_debug.py --as-of 2026-09-04 --out-dir ../tmp
"""

import argparse
import datetime as dt
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill

import refresh_fa_securities as J
from fa import securities as sec
from ta.common import get_supabase_client, safe_execute

V8, V9 = "CTCK_V8", "CTCK_V9_DRAFT"
BLOCKS = {"quality": sec.QUALITY_CRITERIA, "cycle": sec.CYCLE_CRITERIA,
          "valuation": sec.VALUATION_CRITERIA}

# Free-text reasons are for humans; the BA asked for codes so rows can be
# grouped. Mapped from the reason each criterion recorded, longest match first.
REASON_CODES = [
    ("no eligible funding cost", "FUNDING_MISSING"),
    ("downstream of an unusable", "FUNDING_MISSING"),
    ("market share not published", "NO_SOURCE_MARKET_SHARE"),
    ("ATTC", "NO_SOURCE_ATTC"),
    ("mapping not LOCKED", "C18_UNLOCKED"),
    ("formula withdrawn", "C20_WITHDRAWN_V9"),
    ("no core P/E history", "C19_NO_PE_HISTORY"),
    ("no sector distribution", "NO_PEER_DISTRIBUTION"),
    ("needs 8+ quarters", "SHORT_CORE_HISTORY"),
    ("no provision or earning-asset", "NO_PROVISION_DATA"),
    ("reported NPAT <= 0", "SPECIAL_CASE_NEG_NPAT"),
]


def reason_code(reason: str | None) -> str:
    if not reason:
        return ""
    for needle, code in REASON_CODES:
        if needle.lower() in reason.lower():
            return code
    return "OTHER"


def block_totals(row: dict) -> dict:
    """Earned and AVAILABLE max per block, derived from the criterion columns.

    `available_max` is not stored per block — only the total is — so it is
    rebuilt here: a criterion with a null score contributed nothing to the
    denominator. This is the number the UI should be showing beside each block
    instead of the rubric's static maximum.
    """
    out = {}
    for name, keys in BLOCKS.items():
        earned = avail = 0.0
        for k in keys:
            v = row.get(f"{k}_score")
            if v is not None:
                earned += float(v)
                avail += sec.CRITERION_POINTS[k]
        out[f"{name}_earned"] = earned
        out[f"{name}_available_max"] = avail
        out[f"{name}_static_max"] = sum(sec.CRITERION_POINTS[k] for k in keys)
    return out


def fetch(client, symbols, as_of, version) -> dict:
    out = {}
    for s in symbols:
        r = safe_execute(
            client.table("fa_securities_scores").select("*")
            .eq("symbol", s).eq("as_of_date", as_of).eq("model_version", version),
            label=f"{version} {s}").data
        if r:
            out[s] = r[0]
    return out


def recompute_inputs(client, symbols, as_of, quarter, coe):
    """Recompute the C19 chain BOTH ways, to isolate the regression.

    `with_history` is what the daily path produces; `without_history` is what
    the backfill produces, because its call omits the argument and the parameter
    defaults to None. Nothing here is written back — this is a measurement.
    """
    prices = J.latest_prices(client, symbols, as_of)
    out = {}
    for s in symbols:
        st = sec.load_statements(client, s)
        if not st.get("income"):
            continue
        qs, oq, cq = sec.ttm_window(quarter)
        core = sec.compute_core(st, qs, oq, cq)
        hist = J.core_history(st, quarter)
        price = prices.get(s)
        with_h = J.valuation_inputs(st, core, quarter, price, coe, hist)
        without_h = J.valuation_inputs(st, core, quarter, price, coe)   # no history
        bal = st.get("balance", {}).get(cq, {})
        shares = bal.get(sec.BS_SHARES)
        out[s] = {
            "core": core, "history": hist, "price": price, "shares": shares,
            "equity": bal.get(sec.BS_EQUITY),
            "market_cap": (price * shares) if (price and shares) else None,
            "with_history": with_h, "without_history": without_h,
            "pe_history_n": len([h for h in hist[1:]
                                 if h["core_npat"] and h["core_npat"] > 0]),
        }
    return out


HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="333333")
WARN = PatternFill("solid", fgColor="FFF3CD")


def sheet(wb, title, headers, rows, highlight_col=None):
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for c in ws[1]:
        c.font, c.fill = HDR, HDR_FILL
    for r in rows:
        ws.append(r)
    if highlight_col:
        idx = headers.index(highlight_col) + 1
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for c in row:
                if c.value in (None, "", "N/A", "MISSING", "FAIL"):
                    c.fill = WARN
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, 1):
        width = max(len(str(h)) + 2,
                    *(len(str(r[i - 1])) + 2 for r in rows[:200]) if rows else (12,))
        ws.column_dimensions[ws.cell(1, i).column_letter].width = min(width, 46)
    return ws


def main():
    ap = argparse.ArgumentParser(description="Securities rubric debug export")
    ap.add_argument("--as-of", default=None)
    ap.add_argument("--symbols", nargs="*")
    ap.add_argument("--out-dir", default="../tmp")
    args = ap.parse_args()

    c = get_supabase_client()
    as_of = args.as_of
    if not as_of:
        r = safe_execute(client_tbl := c.table("fa_securities_scores")
                         .select("as_of_date").eq("model_version", V9).eq("symbol", "SSI")
                         .order("as_of_date", desc=True).limit(1), label="latest").data
        as_of = r[0]["as_of_date"]
    quarter = J.latest_quarter(c)
    symbols = args.symbols or J.broker_universe(c)
    coe = J.risk_free_rate(c) + J.EQUITY_RISK_PREMIUM

    print(f"=== securities debug export {as_of} (quality quarter {quarter}) ===")
    v9 = fetch(c, symbols, as_of, V9)
    v8 = fetch(c, symbols, as_of, V8)
    symbols = [s for s in symbols if s in v9]
    print(f"  {len(symbols)} brokers with a {V9} row; {len(v8)} with {V8}")
    print(f"  recomputing the C19 chain both ways ...")
    calc = recompute_inputs(c, symbols, as_of, quarter, coe)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- 0. what this file is
    sheet(wb, "0_README", ["Item", "Finding"], [
        ["Export", f"{as_of}, model {V9}, quality quarter {quarter}"],
        ["State", "AS STORED. Nothing was repaired or recomputed into the DB."],
        ["C20 status", "N/A BY DESIGN (V9). Formula withdrawn — it scored 0 for "
                       "all 30 brokers with a reading. Its 12 points leave the denominator."],
        ["C19 status", "N/A BY REGRESSION, unrelated to C20. run_backfill calls "
                       "valuation_inputs() WITHOUT the `history` argument; it defaults "
                       "to None, so the core-P/E history is empty and the percentile "
                       "cannot be formed. The daily path passes it. See 6_Root_cause."],
        ["Cycle 8/30 in the UI", "The UI prints earned/STATIC rubric max. C18 is N/A, "
                                 "so cycle's AVAILABLE max is 23, not 30. Valuation shows "
                                 "0/20 where its available max is 0. See 3_Score_totals, "
                                 "columns *_available_max vs *_static_max."],
        ["normalized_core_npat", "NOT IMPLEMENTED. core_pe uses raw core_npat_ttm; the "
                                 "rubric specifies Normalized_Core_NPAT_V3. Reported as "
                                 "a gap, not silently substituted."],
    ])
    # --- 1. C19 inputs, with the source/status/confidence of each field
    hdr1 = ["symbol", "as_of_date", "reported_npat_ttm", "core_npat_ttm",
            "normalized_core_npat", "shares_outstanding", "market_price", "market_cap",
            "core_pe", "core_npat_source", "core_npat_status", "core_npat_confidence",
            "funding_source", "funding_status", "funding_confidence",
            "price_source", "shares_source", "pe_history_n", "pe_history_required"]
    rows1 = []
    for s in symbols:
        d, fm = calc.get(s, {}), (v9[s].get("field_metadata") or {})
        cn = fm.get("core_npat_ttm", {})
        fc = fm.get("eligible_funding_cost", {})
        rows1.append([
            s, as_of, fm.get("reported_npat_ttm", {}).get("value"), cn.get("value"),
            # Not implemented — reported as blank rather than filled with core_npat,
            # which would hide that the rubric asks for a normalized figure.
            None,
            d.get("shares"), d.get("price"), d.get("market_cap"),
            (d.get("with_history") or {}).get("core_pe"),
            cn.get("source_field"), cn.get("status"), cn.get("confidence"),
            fc.get("source_type"), fc.get("status"), fc.get("confidence"),
            "ta_ohlcv.close", sec.BS_SHARES, d.get("pe_history_n"), 6,
        ])
    sheet(wb, "1_C19_inputs", hdr1, rows1, highlight_col="core_pe")

    # --- 2. C19 and C20 results
    hdr2 = ["symbol", "as_of_date", "c19_score", "c19_status", "c19_reason_code",
            "c19_available_max", "c20_score_production", "c20_status", "c20_reason_code",
            "c20_available_max", "c20_shadow_a_pb_ratio_g2", "c20_shadow_b_pb_percentile",
            "c20_shadow_normalized_total_roe"]
    rows2 = []
    for s in symbols:
        r, flags = v9[s], (v9[s].get("dependency_flags") or {})
        shadow = ((r.get("field_metadata") or {}).get("c20_shadow") or {})
        a = (shadow.get("a_absolute") or {}).get("by_g", {}).get("0.020", {})
        b = shadow.get("b_relative") or {}
        rows2.append([
            s, as_of, r.get("c19_score"),
            "N/A" if r.get("c19_score") is None else "OK",
            reason_code((flags.get("c19") or {}).get("reason")),
            0 if r.get("c19_score") is None else sec.CRITERION_POINTS["c19"],
            r.get("c20_score"), (flags.get("c20") or {}).get("status", "N/A"),
            reason_code((flags.get("c20") or {}).get("reason")),
            0 if r.get("c20_score") is None else sec.CRITERION_POINTS["c20"],
            a.get("pb_ratio"), b.get("pb_hist_percentile"),
            (shadow.get("a_absolute") or {}).get("normalized_total_roe"),
        ])
    sheet(wb, "2_C19_C20_results", hdr2, rows2, highlight_col="c19_score")

    # --- 3. totals, with available vs static max per block
    hdr3 = ["symbol", "as_of_date", "quality_earned", "quality_available_max",
            "quality_static_max", "cycle_earned", "cycle_available_max",
            "cycle_static_max", "valuation_earned", "valuation_available_max",
            "valuation_static_max", "earned_total", "available_max_total", "coverage",
            "provisional_score", "final_fa_score", "validity_status",
            "validity_reason_code", "model_version", "ui_shows_cycle", "ui_shows_valuation"]
    rows3 = []
    for s in symbols:
        r = v9[s]
        b = block_totals(r)
        core_ok = any(r.get(f"{k}_score") is not None for k in ("c1", "c2", "c3"))
        val_ok = any(r.get(f"{k}_score") is not None for k in sec.VALUATION_CRITERIA)
        code = ("" if r["fa_status"] == "PUBLISHABLE" else
                "NO_CORE_EARNINGS" if not core_ok else
                "NO_VALUATION" if not val_ok else
                "COVERAGE_BELOW_50" if r["coverage"] < .5 else "COVERAGE_BELOW_70")
        rows3.append([
            s, as_of, b["quality_earned"], b["quality_available_max"], b["quality_static_max"],
            b["cycle_earned"], b["cycle_available_max"], b["cycle_static_max"],
            b["valuation_earned"], b["valuation_available_max"], b["valuation_static_max"],
            r["earned_score"], r["available_max"], r["coverage"],
            r["normalized_fa_score"], r["normalized_fa_score"], r["fa_status"], code,
            r["model_version"],
            f"{b['cycle_earned']:g}/{b['cycle_static_max']:g}",
            f"{b['valuation_earned']:g}/{b['valuation_static_max']:g}",
        ])
    sheet(wb, "3_Score_totals", hdr3, rows3, highlight_col="validity_reason_code")

    # --- 4. C18
    rows4 = [[s, as_of, v9[s].get("c18_score"),
              "N/A" if v9[s].get("c18_score") is None else "OK",
              0 if v9[s].get("c18_score") is None else sec.CRITERION_POINTS["c18"],
              sec.CRITERION_POINTS["c18"],
              reason_code(((v9[s].get("dependency_flags") or {}).get("c18") or {}).get("reason"))]
             for s in symbols]
    sheet(wb, "4_C18", ["symbol", "as_of_date", "c18_score", "c18_status",
                        "c18_available_max", "c18_static_max", "c18_reason_code"], rows4)

    # --- 5. V8 vs V9, same symbol and date
    hdr5 = ["symbol", "as_of_date", "field", "v8", "v9", "changed"]
    rows5 = []
    for s in symbols:
        if s not in v8:
            continue
        a, b_ = v8[s], v9[s]
        ba, bb = block_totals(a), block_totals(b_)
        pairs = [
            ("normalized_core_npat", None, None),
            ("core_pe", None, None),
            ("c19_score", a.get("c19_score"), b_.get("c19_score")),
            ("c19_status", "OK" if a.get("c19_score") is not None else "N/A",
                           "OK" if b_.get("c19_score") is not None else "N/A"),
            ("c19_available_max", 0 if a.get("c19_score") is None else 8,
                                  0 if b_.get("c19_score") is None else 8),
            ("c20_status", "SCORED" if a.get("c20_score") is not None else "N/A",
                           "WITHDRAWN_V9"),
            ("valuation_available_max", ba["valuation_available_max"], bb["valuation_available_max"]),
            ("available_max_total", a["available_max"], b_["available_max"]),
            ("coverage", a["coverage"], b_["coverage"]),
            ("provisional_score", a["normalized_fa_score"], b_["normalized_fa_score"]),
            ("final_fa_score", a["normalized_fa_score"], b_["normalized_fa_score"]),
            ("fa_status", a["fa_status"], b_["fa_status"]),
        ]
        for name, x, y in pairs:
            rows5.append([s, as_of, name, x, y, "" if x == y else "CHANGED"])
    sheet(wb, "5_V8_vs_V9_diff", hdr5, rows5, highlight_col="changed")

    # --- 6. the root cause, computed both ways
    hdr6 = ["symbol", "pe_history_n", "required", "core_pe_with_history",
            "p_core_pe_with_history", "c19_would_score", "core_pe_without_history",
            "p_core_pe_without_history", "c19_actually_scored", "diagnosis"]
    rows6 = []
    for s in symbols:
        d = calc.get(s, {})
        w, wo = d.get("with_history") or {}, d.get("without_history") or {}
        pw = w.get("p_core_pe")
        would = None
        if pw is not None:
            would = sec._pctile_bands(pw, sec.C19_CORE_PE)
        rows6.append([
            s, d.get("pe_history_n"), 6, w.get("core_pe"), pw, would,
            wo.get("core_pe"), wo.get("p_core_pe"), v9[s].get("c19_score"),
            "history argument missing in run_backfill -> empty PE history -> no percentile"
            if (pw is not None and wo.get("p_core_pe") is None) else
            "insufficient core-PE history even WITH the argument",
        ])
    sheet(wb, "6_Root_cause", hdr6, rows6, highlight_col="c19_actually_scored")

    # --- 7. the code the BA asked to see alongside the numbers
    import inspect
    extracts = [
        ("C19 + C20 production", "fa/securities.py", inspect.getsource(sec.score_valuation)),
        ("C20 shadow A (Gordon)", "fa/securities.py", inspect.getsource(sec.c20_shadow_a)),
        ("C20 shadow A guard", "fa/securities.py", inspect.getsource(sec.justified_pb)),
        ("C20 shadow B (own history)", "fa/securities.py", inspect.getsource(sec.c20_shadow_b)),
        ("Valuation inputs (core_pe, p_core_pe)", "refresh_fa_securities.py",
         inspect.getsource(J.valuation_inputs)),
        ("Shared core history builder", "refresh_fa_securities.py",
         inspect.getsource(J.core_history)),
        ("available_max_total + coverage + validity gate", "fa/securities.py",
         inspect.getsource(sec.assemble)),
        ("DAILY call site — passes history (correct)", "refresh_fa_securities.py",
         "ctx.update(valuation_inputs(st, core, quarter, prices.get(sym), coe, history))"),
        ("BACKFILL call site — omits history (the regression)", "refresh_fa_securities.py",
         "d[\"ctx\"].update(valuation_inputs(d[\"statements\"], d[\"core\"], quarter,\n"
         "                                 price_asof(prices.get(sym, []), as_of), coe))\n"
         "# `history` defaults to None -> hist=[] -> pe_hist=[] -> p_core_pe=None -> C19 N/A"),
        ("API rows returned to the frontend", "dashboard/src/lib/cached-data.ts",
         "getSecRows(date): select * from fa_securities_scores\n"
         "  where as_of_date = date and model_version = SEC_ACTIVE_MODEL\n"
         "    and score_status = 'OFFICIAL'\n"
         "  order by normalized_fa_score desc nulls last, symbol"),
        ("FRONTEND block cell — uses the STATIC max", "sec-scanner-client.tsx",
         "{r[b.key] === null ? '—' : `${Number(r[b.key]).toFixed(0)}/${b.max}`}\n"
         "# b.max is the RUBRIC maximum (quality 50, cycle 30, valuation 20).\n"
         "# It is NOT available_max, so a block whose criteria are N/A still\n"
         "# prints the full denominator: cycle 8/30 where available is 23,\n"
         "# valuation 0/20 where available is 0."),
        ("FRONTEND criterion cell — em dash for N/A", "sec-scanner-client.tsx",
         "v === null ? <span title={dependency_flags[key].reason}>—</span> : `${v}/${c.max}`"),
        ("FRONTEND Final FA Score", "sec-scanner-client.tsx",
         "r.normalized_fa_score === null ? '—' : r.normalized_fa_score.toFixed(1)\n"
         "# = earned_score / available_max * 100, computed server-side in assemble()"),
    ]
    ws = wb.create_sheet("7_Code")
    ws.append(["Area", "File", "Code"])
    for cell in ws[1]:
        cell.font, cell.fill = HDR, HDR_FILL
    for name, path, code in extracts:
        ws.append([name, path, code])
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 120
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    out = Path(args.out_dir) / f"ctck_debug_{as_of.replace('-','')}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"\n  wrote {out} ({len(wb.sheetnames)} sheets, {len(symbols)} brokers)")
    return out


if __name__ == "__main__":
    main()
